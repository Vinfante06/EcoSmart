from django.http import HttpResponse

from .servicios import convertir_a_float


def formato_moneda(valor):
    return f"${convertir_a_float(valor):,.2f}"


def _nombre_categoria(movimiento):
    return str(movimiento.categoria) if movimiento.categoria else "Sin categoría"


def generar_reporte_pdf(datos):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="reporte_{datos["mes"]:02d}_{datos["anio"]}.pdf"'
    )

    documento = SimpleDocTemplate(response, pagesize=letter)
    contenido = []
    estilos = getSampleStyleSheet()
    titulo = ParagraphStyle(
        "TituloEcoSmart",
        parent=estilos["Heading1"],
        fontSize=24,
        textColor=colors.HexColor("#1a472a"),
        spaceAfter=30,
        alignment=1,
    )
    encabezado = ParagraphStyle(
        "EncabezadoEcoSmart",
        parent=estilos["Heading2"],
        fontSize=14,
        textColor=colors.HexColor("#2d7a4a"),
        spaceAfter=12,
        spaceBefore=12,
    )

    def agregar_tabla(filas, anchos, color_encabezado="#2d7a4a", color_fondo="#f5f5f5"):
        tabla = Table(filas, colWidths=anchos)
        tabla.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(color_encabezado)),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 11),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor(color_fondo)),
                    ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                ]
            )
        )
        contenido.append(tabla)

    contenido.append(
        Paragraph(
            f'Reporte Financiero - {datos["mes_nombre"]} {datos["anio"]}',
            titulo,
        )
    )
    contenido.append(Spacer(1, 0.3 * inch))

    contenido.append(Paragraph("Presupuestos vs Gastos", encabezado))
    if datos["presupuestos"]:
        filas = [["Categoría", "Límite", "Gasto", "Disponible", "Estado"]]
        for item in datos["presupuestos"]:
            filas.append(
                [
                    str(item["presupuesto"].categoria),
                    formato_moneda(item["presupuesto"].monto_limite),
                    formato_moneda(item["total_gastos"]),
                    formato_moneda(item["restante"]),
                    "EXCEDIDO" if item["restante"] < 0 else "OK",
                ]
            )
        agregar_tabla(filas, [1.8 * inch, 1.1 * inch, 1.1 * inch, 1.1 * inch, 0.9 * inch])
    else:
        contenido.append(Paragraph("No hay presupuestos registrados para este mes.", estilos["Normal"]))

    contenido.append(Spacer(1, 0.2 * inch))
    contenido.append(Paragraph("Ingresos del Mes", encabezado))
    if datos["ingresos"].exists():
        filas = [["Descripción", "Categoría", "Monto", "Fecha"]]
        for ingreso in datos["ingresos"]:
            filas.append(
                [
                    ingreso.descripcion,
                    _nombre_categoria(ingreso),
                    formato_moneda(ingreso.monto),
                    ingreso.fecha.strftime("%d/%m/%Y"),
                ]
            )
        filas.append(["TOTAL", "", formato_moneda(datos["total_ingresos"]), ""])
        agregar_tabla(filas, [2.2 * inch, 1.5 * inch, 1.3 * inch, 1.2 * inch], "#2d7a4a", "#e8f5e9")
    else:
        contenido.append(Paragraph("No hay ingresos registrados para este mes.", estilos["Normal"]))

    contenido.append(Spacer(1, 0.2 * inch))
    contenido.append(Paragraph("Gastos del Mes", encabezado))
    if datos["gastos"].exists():
        filas = [["Descripción", "Categoría", "Monto", "Fecha"]]
        for gasto in datos["gastos"]:
            filas.append(
                [
                    gasto.descripcion,
                    _nombre_categoria(gasto),
                    formato_moneda(gasto.monto),
                    gasto.fecha.strftime("%d/%m/%Y"),
                ]
            )
        filas.append(["TOTAL", "", formato_moneda(datos["total_gastos"]), ""])
        agregar_tabla(filas, [2.2 * inch, 1.5 * inch, 1.3 * inch, 1.2 * inch], "#c62828", "#ffebee")
    else:
        contenido.append(Paragraph("No hay gastos registrados para este mes.", estilos["Normal"]))

    contenido.append(Spacer(1, 0.2 * inch))
    contenido.append(Paragraph("Resumen Mensual", encabezado))
    agregar_tabla(
        [
            ["Total Ingresos", formato_moneda(datos["total_ingresos"])],
            ["Total Gastos", formato_moneda(datos["total_gastos"])],
            ["Balance", formato_moneda(datos["balance"])],
        ],
        [3 * inch, 2 * inch],
        "#2d7a4a",
        "#f5f5f5",
    )

    contenido.append(PageBreak())
    contenido.append(Paragraph("Objetivos de Ahorro y Progreso", encabezado))
    if datos["objetivos"]:
        filas = [["Objetivo", "Meta", "Ahorrado", "Progreso", "Restante"]]
        for item in datos["objetivos"]:
            objetivo = item["obj"]
            filas.append(
                [
                    objetivo.nombre,
                    formato_moneda(objetivo.monto_objetivo),
                    formato_moneda(objetivo.monto_ahorrado),
                    f'{item["progreso"]}%',
                    formato_moneda(item["restante"]),
                ]
            )
        agregar_tabla(filas, [1.8 * inch, 1.2 * inch, 1.2 * inch, 1.2 * inch, 1.2 * inch], "#2d7a4a", "#f0fdf0")
    else:
        contenido.append(Paragraph("No hay objetivos de ahorro registrados.", estilos["Normal"]))

    documento.build(contenido)
    return response


def generar_reporte_excel(datos):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    titulo_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    titulo_fill = PatternFill(start_color="1a472a", end_color="1a472a", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2d7a4a", end_color="2d7a4a", fill_type="solid")
    total_font = Font(name="Calibri", size=11, bold=True)
    total_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws.merge_cells("A1:E1")
    celda = ws["A1"]
    celda.value = f'Reporte Financiero - {datos["mes_nombre"]} {datos["anio"]}'
    celda.font = titulo_font
    celda.fill = titulo_fill
    celda.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25

    row = 3

    def section_header(label, fill=header_fill):
        nonlocal row
        ws.merge_cells(f"A{row}:E{row}")
        celda_header = ws[f"A{row}"]
        celda_header.value = label
        celda_header.font = header_font
        celda_header.fill = fill
        celda_header.alignment = Alignment(horizontal="left", vertical="center")
        row += 1

    def add_row(values, fmt=None, bold=False):
        nonlocal row
        for col, value in enumerate(values, 1):
            celda_actual = ws.cell(row=row, column=col)
            celda_actual.value = value
            celda_actual.border = border
            celda_actual.alignment = Alignment(horizontal="center")
            if bold:
                celda_actual.font = total_font
                celda_actual.fill = total_fill
            if fmt and fmt.get(col):
                celda_actual.number_format = fmt[col]
        row += 1

    section_header("Presupuestos vs Gastos")
    if datos["presupuestos"]:
        add_row(["Categoría", "Límite", "Gasto", "Disponible", "Estado"], bold=True)
        for item in datos["presupuestos"]:
            add_row(
                [
                    str(item["presupuesto"].categoria),
                    convertir_a_float(item["presupuesto"].monto_limite),
                    convertir_a_float(item["total_gastos"]),
                    convertir_a_float(item["restante"]),
                    "EXCEDIDO" if item["restante"] < 0 else "OK",
                ],
                {2: "$#,##0.00", 3: "$#,##0.00", 4: "$#,##0.00"},
            )
    else:
        ws.merge_cells(f"A{row}:E{row}")
        ws.cell(row=row, column=1).value = "No hay presupuestos registrados para este mes."
        row += 1
    row += 1

    section_header("Ingresos del Mes")
    if datos["ingresos"].exists():
        add_row(["Descripción", "Categoría", "Monto", "Fecha"], bold=True)
        for ingreso in datos["ingresos"]:
            add_row(
                [ingreso.descripcion, _nombre_categoria(ingreso), convertir_a_float(ingreso.monto), ingreso.fecha],
                {3: "$#,##0.00", 4: "dd/mm/yyyy"},
            )
        add_row(["TOTAL", "", convertir_a_float(datos["total_ingresos"]), ""], {3: "$#,##0.00"}, bold=True)
    else:
        ws.merge_cells(f"A{row}:E{row}")
        ws.cell(row=row, column=1).value = "No hay ingresos registrados para este mes."
        row += 1
    row += 1

    section_header(
        "Gastos del Mes",
        PatternFill(start_color="C62828", end_color="C62828", fill_type="solid"),
    )
    if datos["gastos"].exists():
        add_row(["Descripción", "Categoría", "Monto", "Fecha"], bold=True)
        for gasto in datos["gastos"]:
            add_row(
                [gasto.descripcion, _nombre_categoria(gasto), convertir_a_float(gasto.monto), gasto.fecha],
                {3: "$#,##0.00", 4: "dd/mm/yyyy"},
            )
        add_row(["TOTAL", "", convertir_a_float(datos["total_gastos"]), ""], {3: "$#,##0.00"}, bold=True)
    else:
        ws.merge_cells(f"A{row}:E{row}")
        ws.cell(row=row, column=1).value = "No hay gastos registrados para este mes."
        row += 1
    row += 1

    section_header("Resumen Mensual")
    for descripcion, valor in [
        ("Total Ingresos", datos["total_ingresos"]),
        ("Total Gastos", datos["total_gastos"]),
        ("Balance", datos["balance"]),
    ]:
        add_row([descripcion, convertir_a_float(valor)], {2: "$#,##0.00"}, bold=True)
    row += 1

    section_header("Objetivos de Ahorro y Progreso")
    if datos["objetivos"]:
        add_row(["Objetivo", "Meta", "Ahorrado", "Progreso %", "Restante"], bold=True)
        for item in datos["objetivos"]:
            objetivo = item["obj"]
            add_row(
                [
                    objetivo.nombre,
                    convertir_a_float(objetivo.monto_objetivo),
                    convertir_a_float(objetivo.monto_ahorrado),
                    item["progreso"],
                    convertir_a_float(item["restante"]),
                ],
                {2: "$#,##0.00", 3: "$#,##0.00", 4: "0.00", 5: "$#,##0.00"},
            )
    else:
        ws.merge_cells(f"A{row}:E{row}")
        ws.cell(row=row, column=1).value = "No hay objetivos de ahorro registrados."

    for columna, ancho in zip("ABCDE", [25, 15, 15, 15, 15]):
        ws.column_dimensions[columna].width = ancho

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="reporte_{datos["mes"]:02d}_{datos["anio"]}.xlsx"'
    )
    wb.save(response)
    return response
